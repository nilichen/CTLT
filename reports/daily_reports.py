import json
import urllib2
import datetime
import pandas as pd
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
# import csv

import os
import httplib2
from apiclient import discovery
import oauth2client
from oauth2client import client
from oauth2client import tools
import base64
import email

# only change the course_list if needed
course_list = [
            'UBCx__Marketing1x__3T2015',
            # 'UBCx__Climate1x__2T2016',
            'UBCx__SPD1x__2T2016',
            'UBCx__SPD2x__2T2016',
            'UBCx__SPD3x__2T2016',
            'UBCx__UseGen_1x__1T2016',
            'UBCx__UseGen_2x__1T2016',
            'UBCx__PSYC_1x__3T2016',
            'UBCx__PSYC_2x__3T2016',
            'UBCx__PSYC_3x__3T2016',
            'UBCx__PSYC_4x__1T2017',
            'UBCx__PSYC_5x__1T2017',
            'UBCx__PSYC_6x__1T2017',
            'UBCx__ReligionX__1T2017',
            'UBCx__ITSx__3T2016',
]

today = datetime.date.today()
yesterday = today - datetime.timedelta(days=1)
# dictionary to store queried results for all the courses in the course_list
sheets = {}

SCOPES = 'https://www.googleapis.com/auth/gmail.readonly'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'Gmail API Python Quickstart'


def get_credentials():
    """
    Gets valid user credentials from storage.

    If nothing has been stored, or if the stored credentials are invalid,
    the OAuth2 flow is completed to obtain the new credentials.

    Returns:
        Credentials, the obtained credential.
    """
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'gmail-python-quickstart.json')

    store = oauth2client.file.Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print 'Storing credentials to ' + credential_path
    return credentials


def get_content(date):
    """
    Given the date, query the email sent from news@edx.org on that day.

    Returns:
        string of html content on that email, if no email on that day, return an empty string
    """
    tomorrow = date + datetime.timedelta(days=1)
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    service = discovery.build('gmail', 'v1', http=http)
    content = ''
    messages = service.users().messages()\
    .list(userId='me', q="from:news@edx.org after:%s before:%s" 
          % (date.strftime('%Y/%m/%d'), tomorrow.strftime('%Y/%m/%d'))).execute().get('messages', [])
    if messages:
        for message in messages:
            tdata = service.users().messages().get(userId='me', id=message['id'], format='raw').execute()
            msg_str = base64.urlsafe_b64decode(tdata['raw'].encode('ASCII'))
            mime_msg = email.message_from_string(msg_str)
            for part in mime_msg.walk():
                # each part is a either non-multipart, or another multipart message
                # that contains further parts... Message is organized like a tree
                if part.get("content-transfer-encoding"):
                    content = part.get_payload(decode=True) 
    return content

# content = get_content()

def inEmail(content, org='UBCx'):
    """
    Given the string of html content, find the courses that belong to the corresponding organization.

    Returns:
        dictionary {position of the course on the webpage: course_name}
        if no courses are found in the email, return an empty dictionary
    """
    email = {}
    # if content is empty string, skip
    if content:
        pos = 1
        soup = BeautifulSoup(content)
        tds = soup.find_all('td', class_='mcnTextContent')
        for td in tds:
            if td.find('span'):
                if td.find('em'):
                    univ = td.find('em').get_text()
                    # print('span', pos, univ)
                    if org in univ:
                        email[pos] = td.find('span').get_text()
                    pos += 1

            elif td.find('a'):
                course = td.find('a').get_text()
                if '-' in course:
                    if org in course:
                        email[pos] = re.match('(.+)\s-\s.+', course).groups()[0]
                    # print('a', pos, course)
                    pos += 1
    return email


def featured(org='UBCx'):    
    """
    Find courses that belong to the corresponding organization on the featured webpage https://www.edx.org/course
    
    Returns:
        dictionary {position of the featured course: course_name}
        if no featured courses are found, returen an empty dictionary
    """
    feature = {}
    r = urllib2.urlopen('https://www.edx.org/course').read()
    soup = BeautifulSoup(r, 'html.parser')
    courses = soup.find_all('div', class_="course-card verified ")
    pos = 1
    for course in courses:
        if course.find(class_='label').get_text() == org:
            feature[pos] = course.find(class_='title').get_text().strip()
        pos += 1
    return feature


def onHomepage(org='UBCx'):
    """
    Find courses that belong to the corresponding organization on the homepage https://www.edx.org/

    Returns:
        dictionary {position of the course on homepage: course_name}
        if no courses are found on homepage, returen an empty dictionary
    """
    homepage = {}
    url = 'https://www.edx.org/api/discovery/v1/cards?limit=12&tags=14e5ea80-f725-4cd6-82f6-6d2bd63d5159'
    data = json.loads(urllib2.urlopen(url).read())
    pos = 1
    for entry in data:
        if entry['organizations'][0]['display_name'] == org:
            homepage[pos] = entry['title'].strip()
        pos += 1
    return homepage


def appendToExcel(sheets, filepath):
    """
    Iterater over the sheets dictionary, if sheet corresponding to a certain course not present in the workbook, 
    create a new sheet with the queried result, otherwise, append the queried result to an existing sheet.
    Create a workbook if not existed.
    """
    if not os.path.exists(filepath):
        wb = Workbook()
        wb.save(filename = filepath)
    book = load_workbook(filepath)
    writer = pd.ExcelWriter(filepath, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    sheet_names = pd.ExcelFile(filepath).sheet_names
    for sheet_name, df in sheets.iteritems():
        if sheet_name in sheet_names:
            pre = pd.read_excel(filepath, sheet_name)
            # course.Date = course.Date.dt.date       
            startrow = len(pre) + 1
            header = False
        else:
            startrow = 0
            header = True
        df.to_excel(writer, sheet_name, startrow=startrow, header=header, index=False)
        
    writer.save()


def enroll_unenroll_verify(course_list=course_list, start_date=yesterday, end_date=yesterday):
    """
    Iterate over the course_list and query number students enrolled, unenrolled 
    and verified on the day, store the corresponding result in the sheets dictionary.
    """
    for course_id in course_list:
        query = """
        Select Date(time) As Date, Sum(Case When activated Then 1 Else 0 End) As enroll,
        Sum(Case When deactivated Then 1 Else 0 End) As unenroll, 
        Sum(Case When mode='verified' and not activated and mode_changed Then 1 Else 0 End) As verify
        From [{0}.enrollment_events]
        Where Date(time) Between '{1}' And '{2}'
        Group By Date
        Order By Date""".format(course_id, start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))
        df = pd.io.gbq.read_gbq(query, project_id='ubcxdata', verbose=False) 
        df.Date = pd.to_datetime(df.Date, format='%Y-%m-%d').dt.date
        sheets[course_id] = df
        
    # appendToExcel(sheets, df, filepath)


def promote():
    """
    Find promoted courses on the day.

    Returns:
        string 'Homepage/Email/Feature position: course_name'
        if no courses are found, return an empty string
    """
    homepage = onHomepage()
    feature = featured()
    content = get_content(today)
    emaillist = inEmail(content)

    promote = ''
    for k, v in homepage.iteritems():
        promote += 'Homepage %s: %s; ' % (k, v)
    for k, v in emaillist.iteritems():
        promote += 'Email %s: %s; ' % (k, v)
    for k, v in feature.iteritems():
        promote += 'Feature %s: %s; ' % (k, v)

    if promote:
        df = pd.DataFrame({'Date': today, 'Promote': promote}, index=[1])
        df.Date = pd.to_datetime(df.Date, format='%Y-%m-%d').dt.date
        sheets['promote'] = df
        # appendToExcel('promote', df, filepath)



if __name__=="__main__":

    # homepage_courses(filepath='/Users/katrinani/Google Drive/Data scripts/homepage_courses.txt')
    enroll_unenroll_verify()
    promote()
    appendToExcel(sheets, '/Users/katrinani/Google Drive/Data scripts/enroll_unenroll_verify.xlsx')